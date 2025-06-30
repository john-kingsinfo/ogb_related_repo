import mysql.connector
import os
import pandas as pd
import paras
import time

ogbconn = mysql.connector.connect(
    # 連線主機名稱
    host=paras.ogb_host,
    # 登入帳號
    user=paras.ogb_user,
    # 登入密碼
    password=paras.ogb_pwd,
)
ogbc = ogbconn.cursor()


# get all orders/ bundles/ tickets info
member_column = ['Line ID', '姓名', '封鎖Line@', '電話', '註冊時間', '持有員工卡', '持有國寵卡張數', '國寵卡最後到期日', '取得國寵卡消費金額',
                 '雙倍套票購買套數(6萬)', '歡樂套票購買套數(2萬)', '超值套票購買套數(1萬)', '總點數', '已使用點數', '剩餘點數', '有效會員',
                 '入會禮領取', '註冊類型', '推薦者Line ID', '推薦者姓名', '推薦者部門']
member_keys = ['line_id', 'name', 'block', 'phone', 'reg_time', 'group_member', '5', 'vip_end', 'vip_cost',
               '1', '7', '6', 'total_point', 'used_point', 'remaining_point', 'valid_member', 'wel_gift', 'reg_level',
               'referral', 'referral_name', 'referral_dep']
friend_column = ['Line ID', '加好友日期', '封鎖Line@', '加好友場域', '是否註冊會員']
regular_bundle_id_info = {'1': '雙倍寵愛套票60,000', '5': '國寶寵物VIP', '6': '超值寵愛套票10,000', '7': '歡樂寵愛套票20,000'}
# group_member_info = {'4': 'VIP-國寵卡', '9': '國寶員工卡'}

ogb_point_column = ['訂單編號', '訂單成立時間', '會員姓名', 'Line ID', '會員卡', '店名', '銷售代碼', '店員', '應付金額',
                    '實付金額', '發放點數', '使用點數', '點數啟用時間']
bundle_column = ['訂單編號', '訂單成立時間', '會員姓名', 'Line ID', '套票歸屬會員卡', '應付金額', '實付金額', '套票編號',
                 '套票名稱', '銷售代碼', '推薦業務', '販售店家', '會員卡購買日']
ticket_column = ['訂單編號(購入套票的訂單)', '購買日期(購入套票的日期)', '會員姓名', 'Line ID', '編號(票券流水號)', '票券名稱', '套票名稱',
                 '活動名稱', '現在是否過期', '票種名稱', '使用狀態', '是否信託票券', '訂單編號(使用消費的訂單)', '使用於商家', '用途',
                 '面額', '信託金額', '是否為常規套票之票券', '使用時間', '服務時間', '建立時間', '訂單是否取消', '認列金額', 'validity',
                 '票券效期']

members = []
member_card = {4: '國寵卡', 9: '員工卡'}
friends = []
dep_members = {}
give_points = {}
point_order = []
# promo_code = {}
# ignore testing orders
ignored_order = ['GOBO-20240307165542690684', 'GOBO-20240307170539493048']
# ignore _EventOrganizer order
EVENTORGANIZER = '_EventOrganizer'
bundles = []
all_df = []

member_info = {}
min_price_for_points = 100
revoke_bid = 61

start_time = time.time()

# get reward info
sql = '''
SELECT value FROM GoBoGroup.sys_params where id = 1;
'''
ogbc.execute(sql)
min_price_for_points = int(ogbc.fetchone()[0]) * 100

# get line friends info
sql = f'''
SELECT 
    crm.line_userid,
    crm.created_at,
    IF(crm.is_block = 0, '否', '是'),
    inv.description,
    IF(ubi.phone IS NULL, '否', '是')
FROM
    OneGoBoLine.mm_members_business_crm AS crm
        LEFT JOIN
    OneGoBoLine.user_base_info AS ubi ON ubi.line_userid = crm.line_userid AND ubi.business_id = crm.business_id
        LEFT JOIN
    OneGoBoLine.invite_info AS inv ON inv.invite_code = crm.invite_code
WHERE
    crm.business_id IN (1 , {revoke_bid})
order by ubi.business_id desc
;
'''
ogbc.execute(sql)
for r in ogbc.fetchall():
    friends.append(list(r))

# get member department info
sql = f'''
SELECT 
    mdi.line_userid, di.name
FROM
    OneGoBoLine.mm_line_member_dept_info AS mdi
        JOIN
    OneGoBoLine.dept_info AS di ON di.dept_no = mdi.dept_no
;
'''
ogbc.execute(sql)
for r in ogbc.fetchall():
    dep_members[r[0]] = r[1]
######################### 測試用
from openpyxl import load_workbook, Workbook

wb = load_workbook('emp.xlsx')
ws = wb.active
id_dep = {}
for r in range(2, ws.max_row + 1):
    theid = ws.cell(row=r, column=3).value.strip()
    thedep = ws.cell(row=r, column=5).value.strip()
    id_dep[theid] = thedep

sql = '''
select line_userid, idcard from OneGoBoLine.user_base_info where idcard <> '' and idcard is not null;
'''
ogbc.execute(sql)
for r in ogbc.fetchall():
    if r[1] in id_dep:
        dep_members[r[0]] = id_dep[r[1]]

#########################
# to get member basic info
sql = f'''
SELECT 
    ubi.line_userid,
    ubi.name,
    ubi.phone,
    ubi.created_at,
    IF(mm.is_block = 1, '是', '否'),
    IF(ubi.business_id = 1, '是', '否'),
    IF(ubi.idcard is null, '基本註冊', '完整註冊'),
    IF(cer.cloudpos_event_id = 2, '已領', '未領'),
    userbi.line_userid,
    userbi.name
FROM
    OneGoBoLine.user_base_info AS ubi
        LEFT JOIN
    OneGoBoLine.mm_members_business_crm AS mm ON ubi.line_userid = mm.line_userid and ubi.business_id = mm.business_id
        LEFT JOIN
    GoBoCloudPos.cloudpos_event_record AS cer ON ubi.line_userid = cer.line_userid and cer.cloudpos_event_id = 2
        LEFT JOIN
    OneGoBoLine.user_base_info AS userbi ON userbi.referral_code = ubi.upline_referral_code
WHERE
    ubi.business_id in (1, {revoke_bid})
        AND ubi.phone IS NOT NULL
        AND ubi.phone <> ''
;
'''
ogbc.execute(sql)
for r in ogbc.fetchall():
    name = r[1] if r[5] == '是' else '無效會員'
    phone = r[2] if r[5] == '是' else '-----'
    member_info[r[0]] = {'line_id': r[0], 'name': name, 'phone': phone, 'reg_time': r[3], 'block': r[4],
                         'group_member': 'No', '5': 0, 'vip_cost': 0, 'vip_end': '', '1': 0, '7': 0, '6': 0,
                         'total_point': 0, 'used_point': 0, 'remaining_point': 0, 'valid_member': r[5],
                         'reg_level': r[6], 'wel_gift': r[7], 'referral': r[8], 'referral_name': r[9],
                         'referral_dep': dep_members.get(r[8], '')}

# to get summerised points info for all OGB member
sql = '''
SELECT
    line_userid,
    SUM(points),
    SUM(used_points),
    SUM(points) - SUM(used_points)
FROM
    GoBoGroup.points_records
WHERE
    (start_at IS NULL OR start_at < NOW())
        AND method = 1
        AND points_records.invalid = 0
GROUP BY line_userid;
'''
ogbc.execute(sql)
for r in ogbc.fetchall():
    member_info[r[0]].update({'total_point': int(r[1]), 'used_point': int(r[2]), 'remaining_point': int(r[3])})

# to get the paid price for bought bundle (國寵卡相關套票) info from order
sql = f'''
SELECT 
    orr.line_userid,
    tb.ticket_bundle_info_id,
    tb.paid_price
FROM
    GoBoGroup.order_records AS orr
        JOIN
    GoBoGroup.ticket_bundles AS tb ON JSON_CONTAINS(orr.bought_ticket_bundle_ids, cast(tb.id as json), '$')
WHERE
    orr.created_at > '2024-03-8'
        AND tb.canceled_at IS NULL
        AND tb.line_userid <> '{EVENTORGANIZER}'
ORDER BY orr.created_at ASC;
'''
ogbc.execute(sql)
for r in ogbc.fetchall():
    if r[1] == 5:
        member_info[r[0]]['vip_cost'] += r[2]
        continue
    strr1 = str(r[1])
    if strr1 not in member_info[r[0]]:
        member_info[r[0]][strr1] = 0
    member_info[r[0]][strr1] += 1

# to get member card info
sql = '''
SELECT 
    line_userid, group_member_id, MAX(expired_at), count(*)
FROM
    GoBoGroup.one_gobo_member_info
WHERE
    group_member_id IN (4, 9)
        AND state NOT IN (1, 2, 3)
GROUP BY line_userid , group_member_id;
'''
ogbc.execute(sql)
for r in ogbc.fetchall():
    if r[0] not in member_info:
        continue
    if  r[1] == 4:
        member_info[r[0]]['vip_end'] = r[2]
        # 更新國寵卡數量
        member_info[r[0]]['5'] = r[3]
    else:
        # should be group_member
        # if 'del' in r[0]:
        #     continue
        member_info[r[0]]['group_member'] = 'Yes'

for _, v in member_info.items():
    members.append([v[mk] for mk in member_keys])

# to get each order's point info
sql = '''
SELECT order_id, points, start_at FROM GoBoGroup.points_records WHERE method = 1 AND invalid = 0;
'''

ogbc.execute(sql)
for r in ogbc.fetchall():
    give_points[r[0]] = {'points': r[1], 'start_at': r[2]}

# to get order related info
sql = f'''
SELECT 
    ord.order_id,
    ord.created_at,
    ubi.name,
    ubi.line_userid,
    ord.applying_discount_for_consume_points,
    si.name,
    ord.promo_code,
    ord.staff,
    ord.price,
    ord.paid_price,
    IF(ord.specified_points is null, 0, ord.specified_points),
    ord.consume_points
FROM
    GoBoGroup.order_records AS ord
        JOIN
    OneGoBoLine.user_base_info AS ubi ON ubi.line_userid = ord.line_userid
        JOIN
    GoBoGroup.store_info AS si ON si.id = ord.store_id
WHERE
    ord.canceled_at is null
        AND ubi.business_id in (1, {revoke_bid})
        AND (ord.price_for_points >= {min_price_for_points} or ord.consume_points > 0 or ord.specified_points > 0)
ORDER BY ord.created_at ASC;
'''

ogbc.execute(sql)
result = ogbc.fetchall()
for r in result:
    lr = list(r)
    lr.append(r[1])
    if lr[0] in give_points:
        lr[10] = give_points[lr[0]]['points']
        lr[-1] = give_points[lr[0]]['start_at']
    if lr[4] == 0.8:
        if lr[-1].strftime('%Y-%m-%d') > '2024-03-01':
            lr[4] = '國寵卡'
        else:
            lr[4] = '週遊趣'
    elif lr[4] == 0.85:
        lr[4] = '員工卡'
    else:
        lr[4] = ''
    lr[2] = member_info.get(lr[3], {}).get('name', None)
    point_order.append(lr)

# to get pormo code info
# sql = 'SELECT promo_code, name FROM GoBoGroup.promo_code where bid = 1'
# ogbc.execute(sql)
# promo_code = {r[0]: r[1] for r in ogbc.fetchall()}

# to get bought bundle info
sql = f'''
SELECT 
    tb.order_id,
    tb.created_at,
    ubi.name,
    tb.line_userid,
    ogmi.group_member_id,
    tb.price,
    tb.paid_price,
    tb.id,
    tbi.name,
    ord.commission_staff,
    pc.name,
    si.name,
    ogmi.created_at
FROM
    GoBoGroup.ticket_bundles AS tb
        JOIN
    OneGoBoLine.user_base_info AS ubi ON ubi.line_userid = tb.line_userid
        LEFT JOIN
    GoBoGroup.one_gobo_member_info AS ogmi ON ogmi.id = tb.rationed_member_id
        JOIN
    GoBoGroup.order_records AS ord ON ord.order_id = tb.order_id
        JOIN
    GoBoGroup.store_info AS si ON si.id = tb.store_id
        JOIN
    GoBoGroup.ticket_bundle_info AS tbi ON tbi.id = tb.ticket_bundle_info_id
        LEFT JOIN
    GoBoGroup.promo_code AS pc ON pc.promo_code = ord.commission_staff
WHERE
    ubi.business_id in (1, {revoke_bid})
        AND tb.canceled_at IS NULL
        AND tb.line_userid <> '{EVENTORGANIZER}'
        AND ord.canceled_at IS NULL
ORDER BY tb.created_at ASC;
'''
ogbc.execute(sql)
for r in ogbc.fetchall():
    if r[0] in ignored_order:
        continue
    lr = list(r)
    if lr[4]:
        lr[4] = member_card.get(lr[4], f'未設定之會員卡號ID:{lr[4]}')
    # lr[10] = promo_code.get(lr[9], '')
    bundles.append(lr)

# to get tickets created date
sql = '''
SELECT id, transfer_id, created_at FROM GoBoGroup.tickets order by id asc
'''
ogbc.execute(sql)
ticket_bought_date = {}
for r in ogbc.fetchall():
    if r[1]:
        ticket_bought_date[r[0]] = ticket_bought_date[r[1]]
    else:
        ticket_bought_date[r[0]] = r[2]
    

# to get tickets info 
sql = f'''
SELECT
    tic.order_id,
    tic.created_at,
    ubi.name,
    tic.line_userid,
    tic.id,
    ti.name,
    tbi.name,
    oe.name,
    if(tic.expired_at < NOW(), "過期", "有效"),
    si.name,
    if(tic.used_at is null, "未使用", "已使用"),
    if(ti.te_product_id is null, "一般票券", "信託票券"),
    tic.used_order_id,
    sii.name,
    tic.used_for,
    ti.price,
    if(tic.te_price >= 0, tic.te_price, 0),
    tbi.id,
    tic.used_at,
    if(tic.used_at is null, null, if(ord.start_time is null, tic.used_at, ord.start_time)),
    tic.created_at,
    if(ord.canceled_at is null, "否", "是"),
    ti.id,
    tic.transfer_id, tic.validity,
    tic.expired_at

FROM
    GoBoGroup.tickets as tic
        JOIN
    OneGoBoLine.user_base_info as ubi on ubi.line_userid = tic.line_userid
        JOIN
    GoBoGroup.ticket_info as ti on ti.id = tic.ticket_info_id
        LEFT JOIN
    GoBoGroup.ticket_bundle_info as tbi on tbi.id = tic.referred_ticket_bundle_info_id
        LEFT JOIN
    GoBoGroup.order_event as oe on oe.id = tic.referred_event_id
        JOIN
    GoBoGroup.service_info as si on si.id = ti.provider_service_id
        LEFT JOIN
    GoBoGroup.store_info as sii on sii.id = tic.used_store_id
        LEFT JOIN
    GoBoGroup.order_records as ord on ord.order_id = tic.used_order_id
WHERE
    ubi.business_id in (1, {revoke_bid})
        AND tic.used_transfer_id IS NULL
        AND tic.validity = 0
        AND tic.ticket_info_id not in (87, 88, 89)
        AND ((tic.store_id <> 27 AND tic.store_id <> 28) OR tic.store_id IS NULL)
ORDER BY
    tic.id ASC
;
'''
ogbc.execute(sql)
# tickets = [r for r in ogbc.fetchall()]

ticket_bought_idx = 1
user_name_idx = 2
line_id_idx = 3
ticket_id_idx = 4
ti_price_idx = 15
te_price_idx = 16
bundle_id_idx = 17
rec_price_idx = 22
transfer_id_idx = 23
tickets = []
for r in ogbc.fetchall():
    lr = list(r)
    lr[user_name_idx] = member_info.get(lr[line_id_idx], {}).get('name', '不存在的LineID')
    if lr[transfer_id_idx] in ticket_bought_date:
        lr[ticket_bought_idx] = ticket_bought_date[lr[ticket_id_idx]] = ticket_bought_date[lr[transfer_id_idx]]
    if str(lr[bundle_id_idx]) in regular_bundle_id_info.keys():
        lr[bundle_id_idx] = '是'
    else:
        lr[bundle_id_idx] = '否'
    if lr[bundle_id_idx] == '是':
        lr[rec_price_idx] = lr[ti_price_idx] * 0.5
    else:
        if lr[rec_price_idx] in (125, 126):
            lr[rec_price_idx] = 250
        else:
            lr[rec_price_idx] = lr[te_price_idx]
    lr.pop(transfer_id_idx)

    tickets.append(lr)

all_df.append((pd.DataFrame(friends, columns=friend_column), '好友清單'))
all_df.append((pd.DataFrame(members, columns=member_column), '會員資訊'))
all_df.append((pd.DataFrame(point_order, columns=ogb_point_column), '點數資訊'))
all_df.append((pd.DataFrame(bundles, columns=bundle_column), '套票資訊'))
all_df.append((pd.DataFrame(tickets, columns=ticket_column), '票券資訊'))
print(f"It takes {time.time() - start_time :.1f} seconds to process data")
start_time = time.time()
name = '萬國寶會員資訊'
ext = 'xlsx'
fname = f'{name}.{ext}'
sn = 1
while os.path.exists(fname):
    fname = f'{name}_{sn}.{ext}'
    sn += 1
with pd.ExcelWriter(fname) as writer:
    for df in all_df:
        df[0].to_excel(writer, sheet_name=df[1], index=False)
print(f"It takes {time.time() - start_time :.1f} seconds to save excel file")