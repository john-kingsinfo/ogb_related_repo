from datetime import datetime
import math
from openpyxl import load_workbook, Workbook
import os

class stock:
    def __init__(self, name):
        self.name = name

    # excel file name
    fname: str

    # write to file name
    wname: str

    # Pipe/ stack for purchase and used
    # each item in pipe dict format is {itemName: [the Pipe/ stack]]
    # each item in the pipe is [purchaseDate, expiredDate, unitPrice, itemAmount,
    #                            memo, usedAmount, remainingAmount]
    purchasePipe: dict
    usedStack: dict

    # sheets name
    purchaseSheet: str
    usedSheet: str
    stockSheet: str

    # purchase sheet index, 1 based
    pNameIdx: int
    pDateIdx: int
    pAmountIdx: int
    pUnitPriceIdx: int
    pExpDateIdx: int
    pCateIdx: int
    pMemoIdx: int

    # used sheet index, 0 based
    uNameIdx: int
    uDateIdx: int
    uAmountIdx: int
    uCostIdx: int
    uForIdx: int
    uUPriceIdx: int
    uMemoIdx: int

    # pet related index
    uPetIDIdx: int
    uPetNameIdx: int
    uPetOwnerIdx: int

def getDateMonth(theDate):
    sDate = theDate.split('/')
    return f'{sDate[0]}-{sDate[1]}'

def getDate(theDate, withTime=False):
    sDate = theDate.split(' ')
    theDateStr = sDate[0].replace('/', '-')
    if withTime:
        theDateStr += ' ' + sDate[1][:5]
    return theDateStr

def getHisId(petId, theDate):
    # HISID + DATE
    return f'{petId}' + getDate(theDate).replace('-', '')

def roundup(x: float, dig: int):
    b = math.pow(10, dig)
    return math.ceil(x * b) / b

# drog: 藥品
# cDrog: 管制藥品
# mItem: 醫材

cPath = os.path.dirname(os.path.abspath(__file__))
DROG = 'drog'
CDROG = 'cDrog'
MITEM = 'mItem'
CNAME = {DROG: '藥品', CDROG: '管制藥品', MITEM: '衛材用品'}
drog = stock('drog')
cDrog = stock('cDrog')
mItem = stock('mItem')

# item list:
# Purchase Date, Expired Date, Unit Price,
# Item Amount, Memo, Used Amount, Remaining Amount
purchaseInfo = {DROG: {}, CDROG: {}, MITEM: {}}

usedInfo = {DROG: [], CDROG: [], MITEM: []}

# item: DROG: {ITEM-NAME: {RNAME: STR, CATE: STR, YEAR-MONTH: {USED-FOR: INT, PURCHASE-COUNT: INT, TOTAL-COST: INT}}}
stockMonthlyInfo = {DROG: {}, CDROG: {}, MITEM: {}}
stockCost = {}
stockPopRec = {}

# monthly statistic information
# [進貨量, 進貨金額, 使用量, 使用金額, 庫存, 庫存金額] + [藥品] + [衛材]
mStatistic = {}
mOffset = {CDROG: 0, DROG: 6, MITEM: 12}

PCOUNT = '進貨量'
TCOST = 'total-cost'
SCOST = 'sell-cost'
RNAME = '品名'
CATE = '分類/單位'
stockKeys = [PCOUNT, TCOST, SCOST, '門診', '住院', '手術']

FAKESTOCK = '超用量，偽庫存資料。'

DATEFORMAT = '%Y/%m/%d'

# purchase date index
PDIdx = 0
# expired date index
EDIdx = 1
# unit price index
UPIdx = 2
# item amount index
IAIdx = 3
# memo index
MEIdx = 4
# used amount index
UAIdx = 5
# remaining amount index
RAIdx = 6

inputFilePre = '240101-250401'
outputFilePre = '2025-0401'

drog.fname = f'{inputFilePre}_Drugs.xlsx'
drog.wname = f'{outputFilePre}-drog-update.xlsx'
cDrog.fname = f'{inputFilePre}_CtrlDrugs.xlsx'
cDrog.wname = f'{outputFilePre}-cDrog-update.xlsx'
mItem.fname = f'{inputFilePre}_mItem.xlsx'
mItem.wname = f'{outputFilePre}-mItem-update.xlsx'

currentStockName = f'{outputFilePre}-現行庫存.xlsx'

monthlyReportName = f'{outputFilePre}-月總計表.xlsx'
petsListReportName = f'{outputFilePre}-個案總表.xlsx'

drog.purchaseSheet = '進貨記錄'
drog.usedSheet = '流向總表'
drog.stockSheet = '藥品庫存'

drog.purchasePipe = {}
drog.usedStack = {}

drog.pNameIdx = 2
drog.pCateIdx = -1
drog.pDateIdx = 3
drog.pAmountIdx = 4
drog.pUnitPriceIdx = 5
drog.pExpDateIdx = 6
drog.pMemoIdx = 11

drog.uNameIdx = 1
drog.uDateIdx = 2
drog.uForIdx = 3
drog.uAmountIdx = 4
drog.uCostIdx = 10
drog.uUPriceIdx = 11
drog.uMemoIdx = 12

drog.uPetIDIdx = 5
drog.uPetNameIdx = 6
drog.uPetOwnerIdx = 7

cDrog.purchaseSheet = '進貨記錄'
cDrog.usedSheet = '使用總表'
cDrog.stockSheet = '管藥庫存'

cDrog.purchasePipe = {}
cDrog.usedStack = {}

cDrog.pNameIdx = 2
cDrog.pCateIdx = 3
cDrog.pDateIdx = 4
cDrog.pAmountIdx = 6
cDrog.pUnitPriceIdx = 7
cDrog.pExpDateIdx = 5
cDrog.pMemoIdx = 12

cDrog.uNameIdx = 1
cDrog.uDateIdx = 3
cDrog.uForIdx = 4
cDrog.uAmountIdx = 5
cDrog.uCostIdx = 11
cDrog.uUPriceIdx = 13
cDrog.uMemoIdx = 14

cDrog.uPetIDIdx = 6
cDrog.uPetNameIdx = 7
cDrog.uPetOwnerIdx = 8

mItem.purchaseSheet = '進貨記錄'
mItem.usedSheet = '銷貨總表'
mItem.stockSheet = '商品庫存'

mItem.purchasePipe = {}
mItem.usedStack = {}

mItem.pNameIdx = 3
mItem.pCateIdx = 2
mItem.pDateIdx = 4
mItem.pAmountIdx = 5
mItem.pUnitPriceIdx = 6
mItem.pExpDateIdx = 7
mItem.pMemoIdx = 12

mItem.uNameIdx = 2
mItem.uDateIdx = 3
mItem.uForIdx = 4
mItem.uAmountIdx = 5
mItem.uCostIdx = 11
mItem.uUPriceIdx = 12
mItem.uMemoIdx = 13

mItem.uPetIDIdx = 6
mItem.uPetNameIdx = 7
mItem.uPetOwnerIdx = 8


RESTORECOST = '商品回沖不計算成本'
OVERUSEDCOST = '商品超用並以最新成本計算'
NOTFOUNDCOST = '查無品項無法計算成本'

dumpLog = False
targetMapping = [(DROG, drog), (MITEM, mItem), (CDROG, cDrog)]
# targetMapping = [(CDROG, cDrog)]

# the item used history info for each pet
# HIS_ID: {PETNAME: STR, OWNERNAME: STR, HIS: {}}
# HIS: {HIS_ID_WITH_DATE: {DATE: STR, ITEMS: LIST}}
petInfo = {}
PETNAME = 'PETNAME'
OWNER = 'OWNER'
HIS = 'HIS'
THEDATE = 'DATE'
ITEMS = 'ITEMS'

itemRenameList = {'alinamin-F 合力他命F(50.0000 mg/tab)': 'alinamin-F 合力他命Ff(50.0000 mg/tab)',
                  'benazepril（5mg）(5.0000 mg/tab)': 'benazepril（5mg）伏泰康F5(5.0000 mg/tab)',
                  'Plavix （保栓通）(75.0000 mg/tab)': 'Plavix (保栓通) clopidogre 75mg(75.0000 mg/tab)',
                  '平痰息錠  Mubroxol Tablets 30mg(30.0000 mg)': '平痰息錠  Ambroxol 30mg(30.0000 mg)',
                  '全能貓 ＜2.5kg （3劑）(100.0000 0)': '全能貓S＜2.5kg （3劑）(100.0000 0)',
                  '全能貓 2.5-7.5kg （3劑）(100.0000 0)': '全能貓S 2.5-7.5kg （3劑）(100.0000 0)',
                  'Otomax onitment 耳通敏(0.0000 0)': '新耳通敏長效 (MOMETAMAX)(0.0000 0)'}

latest_used_date = '2024/01/01'
# for each type load items info

stockWB = Workbook()
for sheet in stockWB.sheetnames:
    ws = stockWB[sheet]
    stockWB.remove(ws)

allDateMonthList = set()
for tName, target in targetMapping:
    wb = load_workbook(os.path.join(cPath, target.fname))
    dateMonthList = set()
    # load all purchase (in) record, and sort by purchase date(LtoH)->expired date(LtoH)->unit price(HtoL)
    ws = wb[target.purchaseSheet]
    for r in range(2, ws.max_row + 1):
        # init/ load data from excel row data
        name = ws.cell(row=r, column=target.pNameIdx).value
        # check the rename item
        if name in itemRenameList:
            name = itemRenameList[name]
        oname = name
        if target.pCateIdx > 0:
            cate = ws.cell(row=r, column=target.pCateIdx).value.upper()
        else:
            cate = ''
        if tName == CDROG:
            name = name.upper() + ws.cell(row=r, column=target.pCateIdx).value.upper()
        pDate = ws.cell(row=r, column=target.pDateIdx).value
        pAmt = float(ws.cell(row=r, column=target.pAmountIdx).value)
        # tmp = ws.cell(row=r, column=target.pUnitPriceIdx).value
        uPrice = float(ws.cell(row=r, column=target.pUnitPriceIdx).value)
        eDate = ws.cell(row=r, column=target.pExpDateIdx).value
        memo = ws.cell(row=r, column=target.pMemoIdx).value
        if not eDate:
            eDate = ''
        # if pAmt == 0:
        #     continue
        dateMonth = getDateMonth(pDate)
        dateMonthList.add(dateMonth)

        # item list:
        # Purchase Date, Expired Date, Unit Price, Item Amount, Memo, Used Amount, Remaining Amount
        item = [pDate, eDate, uPrice, pAmt, memo, 0, pAmt]
        if name not in target.purchasePipe:
            target.purchasePipe[name] = [item]
        else:
            target.purchasePipe[name].append(item)

        # stockMonthlyInfo used for monthly staticstic
        if name not in stockMonthlyInfo[tName]:
            stockMonthlyInfo[tName][name] = {RNAME: oname, CATE: cate}
        if dateMonth not in stockMonthlyInfo[tName][name]:
            stockMonthlyInfo[tName][name][dateMonth] = {k: 0 for k in stockKeys}

        stockMonthlyInfo[tName][name][dateMonth][PCOUNT] += pAmt
        stockMonthlyInfo[tName][name][dateMonth][TCOST] += roundup(pAmt * uPrice, 4)

    # sort items by Purchase Data(LtH)->Expired Date(LtH)->Unit Price(HtL)
    for name in target.purchasePipe:
        # init used pipe/ stack
        if name not in stockCost:
            stockCost[name] = []
            stockPopRec[name] = []

        items = target.purchasePipe[name]
        # sort purchase item by rules
        target.purchasePipe[name] = sorted(items, key=lambda x: (x[0], x[1], x[2] * -1), reverse=False)
        # append item used index
        target.purchasePipe[name].append(0)

        # add item stock pipe
        for item in target.purchasePipe[name][:-1]:
            # [pAmt, UPrice]
            stockCost[name].append([item[3], item[2]])

    # load all used info
    ws = wb[target.usedSheet]
    # add column name for unit price
    ws.cell(row=1, column=target.uUPriceIdx + 1).value = "平均成本"
    for row in ws.iter_rows(min_row=2):
        usedInfo[tName].append([cell.value for cell in row])

    # sort all used record by used time (LtH), at index 2 in the list
    usedInfo[tName] = sorted(usedInfo[tName], key=lambda x: x[target.uDateIdx], reverse=False)

    # get latest used date
    if usedInfo[tName][-1][target.uDateIdx] > latest_used_date:
        latest_used_date = usedInfo[tName][-1][target.uDateIdx]
    # for each used record do process
    for i in range(len(usedInfo[tName])):
        # append one column for unit price and one column for Memo (list)
        usedInfo[tName][i].append([])
        name = usedInfo[tName][i][target.uNameIdx]
        if tName == CDROG:
            name = name.upper() + usedInfo[tName][i][target.uNameIdx + 1].upper()
        amount = float(usedInfo[tName][i][target.uAmountIdx])
        usedFor = usedInfo[tName][i][target.uForIdx]

        # prepare pet information
        petID = usedInfo[tName][i][target.uPetIDIdx]
        petName = usedInfo[tName][i][target.uPetNameIdx]
        petOwner = usedInfo[tName][i][target.uPetOwnerIdx]
        petHisId = getHisId(petID, usedInfo[tName][i][target.uDateIdx])

        if petID not in petInfo:
            petInfo[petID] = {PETNAME: petName, OWNER: petOwner, HIS: {}}
        if petHisId not in petInfo[petID][HIS]:
            petInfo[petID][HIS][petHisId] = {THEDATE: getDate(usedInfo[tName][i][target.uDateIdx], True),
                                             ITEMS: []}
        # tmp = usedInfo[tName][i]
        petInfo[petID][HIS][petHisId][ITEMS].append([getDate(usedInfo[tName][i][target.uDateIdx], True), usedFor, CNAME[tName], name, amount, 0])

        if amount == 0:
            usedInfo[tName][i][target.uCostIdx] = 0.0
            continue

        # check name format
        if name not in target.purchasePipe:
            found = False
            cname = name + '('
            for dname in target.purchasePipe:
                if dname.startswith(cname):
                    if dumpLog:
                        print(f'@{tName}, {name} should be {dname}')
                    found = True
                    name = dname
                    # update item name for pet info
                    petInfo[petID][HIS][petHisId][ITEMS][-1][3] = name
                    break
            if not found:
                if dumpLog:
                    print(f'@{tName}, {name} not found in Drog Sheet')
                usedInfo[tName][i][target.uCostIdx] = NOTFOUNDCOST
                continue

        # check dateMonth
        dateMonth = getDateMonth(usedInfo[tName][i][target.uDateIdx])
        dateMonthList.add(dateMonth)

        if dateMonth not in stockMonthlyInfo[tName][name]:
            stockMonthlyInfo[tName][name][dateMonth] = {k: 0 for k in stockKeys}
        # add used for amount
        if usedFor not in stockMonthlyInfo[tName][name][dateMonth]:
            if dumpLog:
                print(f'Missing key of {usedFor}')
            stockMonthlyInfo[tName][name][dateMonth][usedFor] = 0
        stockMonthlyInfo[tName][name][dateMonth][usedFor] += amount

        idx = target.purchasePipe[name][-1]
        totalCost = 0
        # restore items?
        # process for floating avg cost
        # if name == 'Alfaxan® 速麻醒(10.0000 mg/ml)':
        #     print(name)
        if amount < 0:
            while amount < 0:
                # 檢查是否回存後，會超出原本該次進貨量，若會繼續往前次進貨回存
                memo = target.purchasePipe[name][idx][MEIdx] if target.purchasePipe[name][idx][MEIdx] is not None else ''
                usedInfo[tName][i][target.uMemoIdx].append(memo)
                if target.purchasePipe[name][idx][UAIdx] + amount < 0:
                    # itemI = target.purchasePipe[name]

                    totalCost -= target.purchasePipe[name][idx][UAIdx] * target.purchasePipe[name][idx][UPIdx]
                    amount += target.purchasePipe[name][idx][UAIdx]
                    # 還原成未使用的狀態
                    target.purchasePipe[name][idx][UAIdx] = 0
                    target.purchasePipe[name][idx][RAIdx] = target.purchasePipe[name][idx][IAIdx]
                    idx -= 1
                    if idx < 0:
                        if dumpLog:
                            print(f'@{tName}, {name} over restore at {usedInfo[tName][i]}')
                        idx = 0
                        # over restore, keep add the cost
                        totalCost += amount * target.purchasePipe[name][0][UPIdx]
                        ########
                        # Stock remaining items into purchasePipe with fist stock info
                        #    # purchase date index
                        #    PDIdx = 0
                        #    # expired date index
                        #    EDIdx = 1
                        #    # unit price index
                        #    UPIdx = 2
                        #    # item amount index
                        #    IAIdx = 3
                        #    # memo index
                        #    MEIdx = 4
                        #    # used amount index
                        #    UAIdx = 5
                        #    # remaining amount index
                        #    RAIdx = 6
                        ########
                        firstItem = list(target.purchasePipe[name][0])
                        firstItem[IAIdx] = abs(amount)
                        firstItem[MEIdx] = '超出庫存原始存量，超存！'
                        firstItem[RAIdx] = abs(amount)
                        target.purchasePipe[name].insert(0, firstItem)
                        amount = 0
                else:
                    target.purchasePipe[name][idx][UAIdx] += amount
                    target.purchasePipe[name][idx][RAIdx] -= amount
                    totalCost += amount * target.purchasePipe[name][idx][UPIdx]
                    amount = 0
            target.purchasePipe[name][-1] = idx
            usedInfo[tName][i][target.uCostIdx] = totalCost #RESTORECOST
        else:
            # amount >= 0
            while amount and idx < len(target.purchasePipe[name]) - 1:
                ra = target.purchasePipe[name][idx][RAIdx]
                memo = target.purchasePipe[name][idx][MEIdx] if target.purchasePipe[name][idx][MEIdx] is not None else ''
                usedInfo[tName][i][target.uMemoIdx].append(memo)
                if usedInfo[tName][i][target.uDateIdx] < target.purchasePipe[name][idx][PDIdx]:
                    print(f'@{tName}, {name} used future stock {usedInfo[tName][i]}')
                if ra <= amount:
                    totalCost += ra * target.purchasePipe[name][idx][UPIdx]
                    target.purchasePipe[name][idx][RAIdx] = 0
                    target.purchasePipe[name][idx][UAIdx] += ra
                    amount -= ra
                    idx += 1
                    target.purchasePipe[name][-1] = idx
                    if idx == len(target.purchasePipe[name]) - 1:
                        print(f"{name} No more stock items, over used")
                        idx -= 1
                        target.purchasePipe[name][-1] = idx
                        break
                else:
                    totalCost += amount * target.purchasePipe[name][idx][UPIdx]
                    target.purchasePipe[name][idx][RAIdx] -= amount
                    target.purchasePipe[name][idx][UAIdx] += amount
                    amount = 0

        # only for over used case
        if amount:
            if dumpLog:
                print(f'@{tName}, {name} over used at {usedInfo[tName][i]}')
            # add cost with latest unit price
            totalCost += amount * target.purchasePipe[name][idx][UPIdx]
            usedInfo[tName][i][target.uMemoIdx].append(OVERUSEDCOST)

            #######
            ## Over used
            #######
            tmp = target.purchasePipe[name]
            overUsedItem = list(target.purchasePipe[name][idx])
            overUsedItem[RAIdx] = amount * -1
            overUsedItem[UAIdx] = amount
            overUsedItem[IAIdx] = 0
            overUsedItem[MEIdx] = FAKESTOCK
            idx += 1
            target.purchasePipe[name].insert(idx, overUsedItem)
            target.purchasePipe[name][-1] = idx
            amount = 0


        usedInfo[tName][i][target.uCostIdx] = float(totalCost)
        if usedInfo[tName][i][target.uAmountIdx] != 0:
            usedInfo[tName][i][target.uUPriceIdx] = roundup(usedInfo[tName][i][target.uCostIdx] / usedInfo[tName][i][target.uAmountIdx], 4)
        # update cost for pet info
        petInfo[petID][HIS][petHisId][ITEMS][-1][-1] = float(totalCost)
        stockMonthlyInfo[tName][name][dateMonth][SCOST] += float(totalCost)

    # get latest column id and add memo column
    usedListLen = len(usedInfo[tName][0])
    ws.cell(row=1, column=target.uMemoIdx + 1).value = '備註'
    # Due to row data sorted, overwrite all rows
    for r in range(2, ws.max_row + 1):
        usedInfo[tName][r - 2][target.uMemoIdx] = ', '.join(usedInfo[tName][r - 2][target.uMemoIdx])
        for i in range(usedListLen):
            ws.cell(row=r, column=i + 1).value = usedInfo[tName][r - 2][i]

    allDateMonthList.update(dateMonthList)
    # stock by month
    #                   0        1        2          3                  4             5          6          7          8           9           10          11            12           13          14
    stockTitles = ['分類/單位', '品名', '進貨', '當月進貨平均成本', '當月進貨總成本', '門診銷量', '住院銷量', '手術銷量', '無銷量來源', '總銷量', '銷貨平均成本', '銷貨總成本', '庫存量', '庫存平均成本', '庫存總成本']
    #              2         3           4        5        6      7     8     9          10          11        12       13           14
    itemKeys = [PCOUNT, '進貨成本', '進貨總成本', '門診', '住院', '手術', '', '總銷', '銷貨平均成本', '總銷成本', '庫存', '庫存成本', '庫存總成本']
    dateMonthList = sorted(list(dateMonthList))
    # 累積到當月的庫存量
    stockByMonth = {}
    for dateMonth in dateMonthList:
        if dateMonth == '2024-11':
            pass
        sheetName = f'{CNAME[tName]}_{dateMonth}_庫存'
        wb.create_sheet(sheetName)
        ws = wb[sheetName]
        itemStockByName = {}
        if dateMonth not in mStatistic:
            mStatistic[dateMonth] = [0] * 18

        # prepare column names
        ws.cell(row=1, column=1).value = f'月份：{dateMonth}'
        for i in range(len(stockTitles)):
            ws.cell(row=2, column=i + 1).value = stockTitles[i]

        # add row at #3
        r = 3
        for name in stockMonthlyInfo[tName].keys():
            item = stockMonthlyInfo[tName][name].get(dateMonth, {})
            ua = usedAmount = item.get('門診', 0) + item.get('住院', 0) + item.get('手術', 0) + item.get('', 0)
            overUsedCost = 0.0
            overUsedCount = 0.0
            if name not in stockByMonth:
                stockByMonth[name] = 0

            if ua >= 0:
                # pop used item from stock pipe
                while ua > 0 and stockCost[name]:
                    if ua >= stockCost[name][0][0]:
                        ua -= stockCost[name][0][0]
                        stockPopRec[name].append(stockCost[name].pop(0))
                        # sc = stockCost[name]
                    else:
                        stockCost[name][0][0] -= ua
                        stockPopRec[name].append([ua, stockCost[name][0][1]])
                        ua = 0
                if ua > 0:
                    overUsedCount = ua
                    print(f'{name} over used {ua} count')
                    if len(stockPopRec[name]) > 0:
                        overUsedCost = ua * stockPopRec[name][-1][1]
                    else:
                        print(stockCost[name], stockPopRec[name])
                        print(f'{name} no used stock??')
            else:
                # ua < 0, restore items
                for bi in range(len(stockPopRec[name]) - 1, -1, -1):
                    if ua + stockPopRec[name][bi][0] <= 0:
                        ua += stockPopRec[name][bi][0]
                        stockCost[name] = [stockPopRec[name][-1]] + stockCost[name]
                        stockPopRec[name] = stockPopRec[name][:-1]
                    else:
                        stockCost[name] = [stockPopRec[name][-1]] + stockCost[name]
                        stockCost[name][0][0] = ua * -1
                        stockPopRec[name][-1][0] += ua
                        ua = 0
                    if ua == 0:
                        break

            if usedAmount != 0:
                item['銷貨平均成本'] = roundup(item[SCOST] / float(usedAmount), 4)
                if usedAmount < 0:
                    item['銷貨平均成本'] *= -1
                item['總銷成本'] = item[SCOST]
            else:
                item['銷貨平均成本'] = 0.0
                item['總銷成本'] = 0.0

            # 累加庫存
            stockByMonth[name] += item.get(PCOUNT, 0) - usedAmount

            # 處理浮點數運算殘值問題
            if stockByMonth[name] != 0 and abs(stockByMonth[name]) < 0.0001:
                print('floating issue!!!!!!!!!!!' + str(stockByMonth[name]))
                stockByMonth[name] = 0

            # 載入當下的庫存量
            sCount = stockByMonth[name]
            # 計算累加庫存成本
            sCost = 0.0
            if sCount >= 0:
                si = 0
                while sCount > 0 and si < len(stockCost[name]):
                    # s = stockCost[name]
                    if sCount >= stockCost[name][si][0]:
                        sCost += stockCost[name][si][0] * stockCost[name][si][1]
                        sCount -= stockCost[name][si][0]
                        si += 1
                    else:
                        sCost += sCount * stockCost[name][si][1]
                        sCount = 0.0
                if sCount > 0:
                    tmp = stockCost[name]
                    print(f'Over stocked {name} with {sCount} number')
                    if len(stockCost[name]) > 0:
                        sCost += sCount * stockCost[name][-1][1]
                    else:
                        sCost += sCount * stockPopRec[name][-1][1]
            else:
                # restore process will put used item from stockPopRec to stockCost
                print(f'Over used {name} with {sCount} number')
                # tmp = stockPopRec[name]
                # tmp1 = stockCost[name]
                # check used stock, if no data, get price from stockCost
                if len(stockPopRec[name]) == 0:
                    pp = stockCost[name][0][1]
                else:
                    pp = stockPopRec[name][-1][1]
                sCost = sCount * pp
                sCount = 0

            sAvgCost = 0.0
            if stockByMonth[name] != 0:
                sAvgCost = roundup(sCost / stockByMonth[name], 4)
                if stockByMonth[name] < 0:
                    sAvgCost *= -1
            itemData = [stockMonthlyInfo[tName][name][CATE], stockMonthlyInfo[tName][name][RNAME]] + [item.get(k, 0) for k in itemKeys]
            #                2         3           4        5        6      7     8     9          10          11        12       13           14
            # itemKeys = [PCOUNT, '進貨成本', '進貨總成本', '門診', '住院', '手術', '', '總銷', '銷貨平均成本', '總銷成本', '庫存', '庫存成本', '庫存總成本']
            # itemData[1] = name
            itemData[3] = 0.0 if item.get(PCOUNT, 0) == 0 else roundup(float(item[TCOST]) / float(item[PCOUNT]), 4)
            itemData[4] = item.get(TCOST, 0)
            itemBase = 9
            itemData[itemBase] = usedAmount
            itemData[itemBase + 1] = item['銷貨平均成本']
            itemData[itemBase + 2] = item['總銷成本']
            itemData[itemBase + 3] = stockByMonth[name]
            itemData[itemBase + 4] = sAvgCost
            itemData[itemBase + 5] = sCost
            for i in range(len(itemData)):
                ws.cell(row=r, column=i + 1).value = itemData[i]
            r += 1

            # 月總表記錄
            ofs = mOffset[tName]
            mStatistic[dateMonth][ofs] += item.get(PCOUNT, 0)
            mStatistic[dateMonth][ofs + 1] += item.get(TCOST, 0)
            mStatistic[dateMonth][ofs + 2] += usedAmount
            mStatistic[dateMonth][ofs + 3] += item.get(SCOST, 0)
            mStatistic[dateMonth][ofs + 4] += stockByMonth[name]
            mStatistic[dateMonth][ofs + 5] += sCost

    wb.save(os.path.join(cPath, target.wname))

    # write back current purchase item stock information
    # item list:
    # Purchase Date, Expired Date, Unit Price, Item Amount, Memo, Used Amount, Remaining Amount
    stockWB.create_sheet(tName)
    stockWS = stockWB[tName]
    stockColumns = ['品名', '庫存數量', '單價', '進貨日期', '到期日期', '備註']
    for c in range(len(stockColumns)):
        stockWS.cell(row=1, column=c + 1).value = stockColumns[c]
    r = 2
    remain = 0
    for name in target.purchasePipe:
        # p = target.purchasePipe[name]
        # if r > 2:
        #     stockWS.cell(row=r - 1, column=7).value = remain
        #     remain = 0
        for ci in range(target.purchasePipe[name][-1], len(target.purchasePipe[name]) - 1):
            if target.purchasePipe[name][ci][6] == 0 and target.purchasePipe[name][ci][4] == FAKESTOCK:
                continue
            stockWS.cell(row=r, column=1).value = name
            stockWS.cell(row=r, column=2).value = target.purchasePipe[name][ci][6]
            stockWS.cell(row=r, column=3).value = target.purchasePipe[name][ci][2]
            stockWS.cell(row=r, column=4).value = target.purchasePipe[name][ci][0]
            stockWS.cell(row=r, column=5).value = target.purchasePipe[name][ci][1]
            stockWS.cell(row=r, column=6).value = target.purchasePipe[name][ci][4]
            remain += target.purchasePipe[name][ci][6]
            r += 1

stockWB.save(os.path.join(cPath, currentStockName))

wb = Workbook()
fp = os.path.join(cPath, monthlyReportName)
sheetName = '113年月總計表'
mColumns = ['類別', '管制藥品進貨量', '管制藥品進貨量金額', '管制藥品使用量', '管制藥品使用量金額', '管制藥品系統庫存量', '管制藥品庫存量系統金額',
            '藥品進貨量', '藥品進貨量金額', '藥品使用量', '藥品使用量金額', '品系統庫存量', '藥品庫存量系統金額',
            '衛材進貨量', '衛材進貨量金額', '衛材使用量', '衛材使用量金額', '衛材系統庫存量', '衛材庫存量系統金額']

for sheet in wb.sheetnames:
    ws = wb[sheet]
    wb.remove(ws)
pYear = ''

yearSum = [0] * 18
r = 2
allDateMonthList = sorted(list(allDateMonthList))
for dateMonth in allDateMonthList:
    year, month = dateMonth.split('-')
    if year != pYear:
        pYear = year
        sheetName = f'{year}年月總計表'
        wb.create_sheet(sheetName)
        ws = wb[sheetName]
        for i in range(len(mColumns)):
            ws.cell(row=1, column=i + 1).value = mColumns[i]
        r = 2
        yearSum = [0] * 18
    ws.cell(row=r, column=1).value = dateMonth
    for c in range(len(mStatistic[dateMonth])):
        ws.cell(row=r, column=c + 2).value = mStatistic[dateMonth][c]
        yearSum[c] += mStatistic[dateMonth][c]
    r += 1
    if month == '12':
        ws.cell(row=r, column=1).value = '總合'
        for c in range(len(yearSum)):
            ws.cell(row=r, column=c + 2).value = yearSum[c]

# if not end in DEC
if month != '12':
    ws.cell(row=r, column=1).value = '總合'
    for c in range(len(yearSum)):
        ws.cell(row=r, column=c + 2).value = yearSum[c]

wb.save(fp)

# the item used history info for each pet
# HIS_ID: {PETNAME: STR, OWNERNAME: STR, HIS: {}}
# HIS: {HIS_ID_WITH_DATE: {DATE: STR, ITEMS: LIST}}
# petInfo = {}
# PETNAME = 'PETNAME'
# OWNER = 'OWNER'
# HIS = 'HIS'
# THEDATE = 'DATE'
# ITEMS = 'ITEMS'

wb = Workbook()
fp = os.path.join(cPath, petsListReportName)
sheetName = '寵物病歷記錄'
# 序號	項目	藥品/衛材/管樂	來源	用量	成本價格	金額	編號	銷貨時間	寵物名	飼主名

petInfoColumn = ['病歷號', '寵物名字', '飼主名字', '看診編號', '看診日期', '來源', '品項名稱', '使用量', '成本']
petInfoColumn = ['序號', '項目', '藥品/衛材/管藥 內容', '來源', '用量', '成本價格', '金額', '編號', '銷貨日期', '寵物名', '飼主名', '寵物病歷號']

petStatisticColumn = ['項次', '成本結算', '病歷號碼', '寵物名', '飼主名', '看診日', '衛材用品', '管制藥品', '一般藥品', '小計']

# wb.save(fp)
# wb = load_workbook(os.path.join(cPath, 'petHis.xlsx'))
for sheet in wb.sheetnames:
    ws = wb[sheet]
    wb.remove(ws)


for dateMonth in allDateMonthList:
    sheetName = f'{dateMonth}'
    r = 2
    wb.create_sheet(sheetName)
    ws = wb[sheetName]
    for i in range(len(petInfoColumn)):
        ws.cell(row=1, column=i + 1).value = petInfoColumn[i]

    r_s = 2
    sheetName = f'{dateMonth}-統計'
    wb.create_sheet(sheetName)
    ws_s = wb[sheetName]
    for i in range(len(petStatisticColumn)):
        ws_s.cell(row=1, column=i + 1).value = petStatisticColumn[i]

    for petId, pInfo in petInfo.items():
        for h in pInfo[HIS]:
            if dateMonth != pInfo[HIS][h][THEDATE][:7]:
                continue
            # ws = wb[dateMonth]
            petStat = {CNAME[c]: 0 for c in CNAME}
            petStat['total'] = 0
            for item in pInfo[HIS][h][ITEMS]:
                # item values: datetime, usedFor, tName, name, amount, cost
                avg = 0.0
                if item[4] != 0:
                    avg = roundup(item[5] / float(item[4]), 4)
                pet = [r - 1, item[2], item[3], item[1], item[4], avg, item[5],
                    h, item[0], pInfo[PETNAME], pInfo[OWNER], petId]
                for c in range(len(pet)):
                    ws.cell(row=r, column=c + 1).value = pet[c]
                r += 1
                petStat[item[2]] += item[5]
                petStat['total'] += item[5]
            ps = [r_s - 1, petStat['total'], petId, pInfo[PETNAME], pInfo[OWNER], pInfo[HIS][h][THEDATE],
                  petStat[CNAME[MITEM]], petStat[CNAME[CDROG]], petStat[CNAME[DROG]], petStat['total']]
            for c in range(len(ps)):
                ws_s.cell(row=r_s, column=c + 1).value = ps[c]
            r_s += 1


wb.save(fp)
